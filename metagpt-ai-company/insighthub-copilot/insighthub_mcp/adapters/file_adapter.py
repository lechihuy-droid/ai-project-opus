"""File-backed adapters for sample Jira, chat, and GitHub exports."""

from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .base import SourceAdapter


def _rows_by_header(path: Path, sheet: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[sheet]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() for value in rows[0]]
    return [
        {headers[index]: value for index, value in enumerate(row)}
        for row in rows[1:]
        if any(value is not None for value in row)
    ]


def _to_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.fromisoformat(str(value)).date()


def _to_datetime(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    return datetime.fromisoformat(str(value))


def _in_period(value: datetime, period_start: date, period_end: date) -> bool:
    return period_start <= value.date() <= period_end


class JiraFileAdapter(SourceAdapter):
    """Read Jira issues and sprint metrics from the sample workbook."""

    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)

    def fetch(self, period_start: date, period_end: date) -> list[dict]:
        issues = []
        for row in _rows_by_header(self.path, "Issues"):
            labels = row.get("Labels") or ""
            issues.append(
                {
                    "key": row["Key"],
                    "summary": row["Summary"],
                    "issue_type": row.get("Type") or "Task",
                    "status": row.get("Status") or "To Do",
                    "priority": row.get("Priority") or "Medium",
                    "assignee": row.get("Assignee") or "",
                    "story_points": float(row.get("Story Points") or 0),
                    "created": _to_date(row.get("Created")).isoformat(),
                    "updated": _to_date(row.get("Updated")).isoformat(),
                    "resolved": (
                        _to_date(row.get("Resolved")).isoformat()
                        if _to_date(row.get("Resolved"))
                        else None
                    ),
                    "sprint": row.get("Sprint") or "",
                    "labels": [item.strip() for item in str(labels).split(",") if item.strip()],
                    "reopened": str(row.get("Reopened") or "").strip().lower() == "yes",
                }
            )

        sprints = []
        for row in _rows_by_header(self.path, "Sprints"):
            committed = float(row.get("Committed SP") or 0)
            completed = float(row.get("Completed SP") or 0)
            sprints.append(
                {
                    "sprint": row["Sprint"],
                    "start": _to_date(row.get("Start")).isoformat(),
                    "end": _to_date(row.get("End")).isoformat(),
                    "committed_sp": committed,
                    "completed_sp": completed,
                    "completion_pct": (completed / committed * 100) if committed else 0.0,
                }
            )

        return [{"issues": issues, "sprints": sprints}]


class ChatFileAdapter(SourceAdapter):
    """Read Slack messages from the sample JSON export."""

    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)

    def fetch(self, period_start: date, period_end: date) -> list[dict]:
        payload = json.loads(self.path.read_text(encoding="utf-8"))
        messages = []
        for message in payload.get("messages", []):
            ts = _to_datetime(message["ts"])
            if _in_period(ts, period_start, period_end):
                messages.append(
                    {
                        "msg_id": message["id"],
                        "channel": message.get("channel") or payload.get("channel") or "",
                        "user": message.get("user") or "",
                        "ts": ts.isoformat(),
                        "text": message.get("text") or "",
                    }
                )
        return messages


class GithubFileAdapter(SourceAdapter):
    """Read commits and pull requests from the sample GitHub export."""

    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)

    def fetch(self, period_start: date, period_end: date) -> list[dict]:
        payload = json.loads(self.path.read_text(encoding="utf-8"))
        commits = []
        for commit in payload.get("commits", []):
            committed_at = _to_datetime(commit["date"])
            if _in_period(committed_at, period_start, period_end):
                commits.append(
                    {
                        "sha": commit["sha"],
                        "message": commit.get("message") or "",
                        "author": commit.get("author") or "",
                        "date": committed_at.isoformat(),
                    }
                )

        prs = []
        for pr in payload.get("pull_requests", []):
            created_at = _to_datetime(pr["created_at"])
            if _in_period(created_at, period_start, period_end):
                merged_at = pr.get("merged_at")
                prs.append(
                    {
                        "number": pr["number"],
                        "title": pr.get("title") or "",
                        "author": pr.get("author") or "",
                        "reviewers": pr.get("reviewers") or [],
                        "state": pr.get("state") or "open",
                        "created": created_at.isoformat(),
                        "merged_at": _to_datetime(merged_at).isoformat() if merged_at else None,
                        "ci_status": pr.get("ci_status") or "",
                    }
                )

        return [{"commits": commits, "pull_requests": prs}]

