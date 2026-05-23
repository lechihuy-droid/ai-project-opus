"""Load InsightHub sample data into a ProjectState."""

from __future__ import annotations

import asyncio
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import yaml
from docx import Document
from fastmcp import Client
from openpyxl import load_workbook

from insighthub.schema import (
    ChatMessage,
    Commit,
    Issue,
    MinuteItem,
    ProjectState,
    PullRequest,
    SourceRef,
    SprintMetric,
    WBSTask,
)
from insighthub_mcp.server import mcp

JIRA_KEY_RE = re.compile(r"[A-Z][A-Z0-9]+-\d+")


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[1]


def _resolve_path(value: str) -> Path:
    path = Path(value)
    if not path.is_absolute():
        path = _repo_root() / path
    return path


def _to_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value))


def _rows_by_header(path: Path, sheet: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[sheet]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = [str(value).strip() for value in rows[0]]
    return [
        {headers[index]: value for index, value in enumerate(row)}
        for row in rows[1:]
        if any(value is not None for value in row)
    ]


def _tool_payload(result: Any) -> dict:
    if hasattr(result, "structured_content") and result.structured_content is not None:
        return result.structured_content
    if hasattr(result, "data") and result.data is not None:
        return result.data
    if hasattr(result, "content") and result.content:
        content = result.content[0]
        text = getattr(content, "text", None)
        if text:
            return json.loads(text)
    if isinstance(result, dict):
        return result
    raise TypeError(f"Unsupported MCP result: {type(result)!r}")


async def _fetch_mcp(period_start: date, period_end: date, connections_path: str) -> tuple[dict, dict, dict]:
    async with Client(mcp) as client:
        start = period_start.isoformat()
        end = period_end.isoformat()
        args = {"period_start": start, "period_end": end, "connections_path": connections_path}
        jira = _tool_payload(await client.call_tool("list_jira_issues", args))
        chat = _tool_payload(await client.call_tool("list_chat_messages", args))
        github = _tool_payload(await client.call_tool("list_code_activity", args))
        return jira, chat, github


def _source_url(base_urls: dict[str, str], system: str, ref_id: str) -> str | None:
    base = (base_urls.get(system) or "").rstrip("/")
    if not base:
        return None
    if system == "jira":
        return f"{base}/{ref_id}"
    if system == "github":
        if ref_id.startswith("PR-"):
            return f"{base}/pull/{ref_id.removeprefix('PR-')}"
        return f"{base}/commit/{ref_id}"
    return f"{base}/{ref_id}"


def _jira_issues(payload: dict, base_urls: dict[str, str]) -> list[Issue]:
    return [
        Issue(
            **item,
            source_ref=SourceRef(
                system="jira",
                ref_id=item["key"],
                label=item["summary"],
                url=_source_url(base_urls, "jira", item["key"]),
            ),
        )
        for item in payload["issues"]
    ]


def _sprints(payload: dict) -> list[SprintMetric]:
    return [
        SprintMetric(
            sprint=item["sprint"],
            committed_sp=item["committed_sp"],
            completed_sp=item["completed_sp"],
            completion_pct=item["completion_pct"],
        )
        for item in payload["sprints"]
    ]


def _messages(payload: dict, base_urls: dict[str, str]) -> list[ChatMessage]:
    return [
        ChatMessage(
            **item,
            source_ref=SourceRef(
                system="slack",
                ref_id=item["msg_id"],
                label=item["text"][:80],
                url=_source_url(base_urls, "slack", item["msg_id"]),
            ),
        )
        for item in payload["messages"]
    ]


def _commits(payload: dict, base_urls: dict[str, str]) -> list[Commit]:
    commits = []
    for item in payload["commits"]:
        ref_id = item["sha"][:7]
        commits.append(
            Commit(
                **item,
                jira_keys=JIRA_KEY_RE.findall(item["message"]),
                source_ref=SourceRef(
                    system="github",
                    ref_id=ref_id,
                    label=item["message"],
                    url=_source_url(base_urls, "github", ref_id),
                ),
            )
        )
    return commits


def _pull_requests(payload: dict, base_urls: dict[str, str]) -> list[PullRequest]:
    prs = []
    for item in payload["pull_requests"]:
        ref_id = f"PR-{item['number']}"
        prs.append(
            PullRequest(
                **item,
                jira_keys=JIRA_KEY_RE.findall(item["title"]),
                source_ref=SourceRef(
                    system="github",
                    ref_id=ref_id,
                    label=item["title"],
                    url=_source_url(base_urls, "github", ref_id),
                ),
            )
        )
    return prs


def _wbs_tasks(path: Path) -> list[WBSTask]:
    tasks = []
    for index, row in enumerate(_rows_by_header(path, "WBS"), start=2):
        task_id = str(row["Task ID"])
        name = str(row["Task Name"])
        tasks.append(
            WBSTask(
                task_id=task_id,
                phase=str(row["Phase"]),
                name=name,
                planned_start=_to_date(row["Planned Start"]),
                planned_end=_to_date(row["Planned End"]),
                planned_mm=float(row.get("Planned MM") or 0),
                planned_pct=float(row.get("Planned %") or 0),
                owner=str(row.get("Owner") or ""),
                jira_key=str(row.get("Jira Key") or ""),
                source_ref=SourceRef(system="wbs", ref_id=task_id or f"row-{index}", label=name),
            )
        )
    return tasks


def _kind_from_heading(text: str) -> str | None:
    return {
        "Decisions": "decision",
        "Action Items": "action_item",
        "Questions": "question",
    }.get(text)


def _due_from_text(text: str) -> date | None:
    match = re.search(r"\b(20\d{2}-\d{2}-\d{2})\b", text)
    return date.fromisoformat(match.group(1)) if match else None


def _owner_from_action(text: str) -> str:
    match = re.match(r"([A-Z][A-Za-z]+)\s+to\s+", text)
    return match.group(1) if match else ""


def _minutes_items(minutes_dir: Path) -> list[MinuteItem]:
    items = []
    for path in sorted(minutes_dir.glob("*.docx")):
        document = Document(path)
        meeting_date = None
        current_kind = None
        kind_counts: dict[str, int] = {}

        for paragraph in document.paragraphs:
            text = paragraph.text.strip()
            if not text:
                continue
            if text.startswith("Date:"):
                meeting_date = date.fromisoformat(text.split(":", 1)[1].strip())
                continue
            heading_kind = _kind_from_heading(text)
            if heading_kind:
                current_kind = heading_kind
                continue
            if current_kind and paragraph.style.name.startswith("List"):
                item_date = meeting_date or date.fromisoformat(path.stem.split("_")[1])
                kind_counts[current_kind] = kind_counts.get(current_kind, 0) + 1
                ref_id = f"{path.name}#{current_kind}-{kind_counts[current_kind]}"
                items.append(
                    MinuteItem(
                        meeting_date=item_date,
                        kind=current_kind,
                        text=text,
                        owner=_owner_from_action(text) if current_kind == "action_item" else "",
                        due=_due_from_text(text),
                        source_ref=SourceRef(system="minutes", ref_id=ref_id, label=text[:80]),
                    )
                )
    return items


def load(connections_path: str = "connections.yaml") -> ProjectState:
    """Load all configured sources into a ProjectState."""

    config_path = _resolve_path(connections_path)
    config = yaml.safe_load(config_path.read_text(encoding="utf-8"))
    period_start = date.fromisoformat(str(config["period"]["start"]))
    period_end = date.fromisoformat(str(config["period"]["end"]))

    base_urls = config.get("base_urls") or {}
    jira_payload, chat_payload, github_payload = asyncio.run(_fetch_mcp(period_start, period_end, connections_path))
    uploads = config["uploads"]

    return ProjectState(
        project_name=config["project"],
        period_start=period_start,
        period_end=period_end,
        issues=_jira_issues(jira_payload, base_urls),
        wbs_tasks=_wbs_tasks(_resolve_path(uploads["wbs"])),
        messages=_messages(chat_payload, base_urls),
        commits=_commits(github_payload, base_urls),
        prs=_pull_requests(github_payload, base_urls),
        minute_items=_minutes_items(_resolve_path(uploads["minutes_dir"])),
        sprints=_sprints(jira_payload),
    )


def load_previous_reports(connections_path: str = "connections.yaml") -> list[str]:
    """Load previous Markdown reports for LLM tone reference only."""

    config_path = _resolve_path(connections_path)
    config = yaml.safe_load(config_path.read_text(encoding="utf-8"))
    previous_dir = _resolve_path(
        str(config.get("previous_reports_dir") or "data/sample/previous_reports")
    )
    if not previous_dir.exists():
        return []
    return [
        path.read_text(encoding="utf-8")
        for path in sorted(previous_dir.glob("*.md"))
        if path.is_file()
    ]


if __name__ == "__main__":
    state = load()
    print(state.source_count())
    print(f"len(state.sprints)={len(state.sprints)}")
