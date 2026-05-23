from __future__ import annotations

import os
import sys

PY311 = r"C:\Users\HUY\AppData\Local\Programs\Python\Python311\python.exe"
if os.path.exists(PY311) and os.environ.get("INSIGHTHUB_PY311_REEXEC") != "1":
    os.environ["INSIGHTHUB_PY311_REEXEC"] = "1"
    os.environ.pop("PYTHONHOME", None)
    os.environ.pop("PYTHONPATH", None)
    os.execv(PY311, [PY311, *sys.argv])

import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any


REPO_ROOT = Path(__file__).resolve().parents[1]
JIRA_KEY_RE = re.compile(r"[A-Z][A-Z0-9]+-\d+")


def _ensure_deps() -> None:
    try:
        import docx  # noqa: F401
        import openpyxl  # noqa: F401
        import yaml  # noqa: F401
    except ModuleNotFoundError:
        raise


_ensure_deps()

import yaml
from docx import Document
from openpyxl import load_workbook

sys.path.insert(0, str(REPO_ROOT))

from insighthub.schema import (  # noqa: E402
    ChatMessage,
    Commit,
    Issue,
    MinuteItem,
    ProjectState,
    PullRequest,
    SourceRef,
    SprintMetric,
    WBSTask,
)


def as_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value))


def as_datetime(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    return datetime.fromisoformat(str(value))


def rows_as_dicts(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(header) for header in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]


def split_labels(value: Any) -> list[str]:
    if value in (None, ""):
        return []
    return [label.strip() for label in str(value).split(",") if label.strip()]


def jira_keys(text: str) -> list[str]:
    return sorted(set(JIRA_KEY_RE.findall(text)))


def load_config() -> dict[str, Any]:
    return yaml.safe_load((REPO_ROOT / "connections.yaml").read_text(encoding="utf-8"))


def load_issues(path: Path) -> list[Issue]:
    issues: list[Issue] = []
    for row in rows_as_dicts(path, "Issues"):
        key = str(row["Key"])
        summary = str(row["Summary"])
        issues.append(
            Issue(
                key=key,
                summary=summary,
                issue_type=str(row["Type"]),
                status=str(row["Status"]),
                priority=str(row["Priority"]),
                assignee=str(row["Assignee"] or ""),
                story_points=float(row["Story Points"] or 0.0),
                created=as_date(row["Created"]),
                updated=as_date(row["Updated"]),
                resolved=as_date(row["Resolved"]),
                sprint=str(row["Sprint"] or ""),
                labels=split_labels(row["Labels"]),
                reopened=str(row["Reopened"]).strip().lower() == "yes",
                source_ref=SourceRef(system="jira", ref_id=key, label=summary),
            )
        )
    return issues


def load_sprints(path: Path) -> list[SprintMetric]:
    sprints: list[SprintMetric] = []
    for row in rows_as_dicts(path, "Sprints"):
        committed = float(row["Committed SP"] or 0.0)
        completed = float(row["Completed SP"] or 0.0)
        completion_pct = completed / committed * 100 if committed else 0.0
        sprints.append(
            SprintMetric(
                sprint=str(row["Sprint"]),
                committed_sp=committed,
                completed_sp=completed,
                completion_pct=completion_pct,
            )
        )
    return sprints


def load_wbs(path: Path) -> list[WBSTask]:
    tasks: list[WBSTask] = []
    for row_num, row in enumerate(rows_as_dicts(path, "WBS"), start=2):
        task_id = str(row["Task ID"])
        name = str(row["Task Name"])
        tasks.append(
            WBSTask(
                task_id=task_id,
                phase=str(row["Phase"]),
                name=name,
                planned_start=as_date(row["Planned Start"]),
                planned_end=as_date(row["Planned End"]),
                planned_mm=float(row["Planned MM"] or 0.0),
                planned_pct=float(row["Planned %"] or 0.0),
                owner=str(row["Owner"] or ""),
                jira_key=str(row["Jira Key"] or ""),
                source_ref=SourceRef(system="wbs", ref_id=task_id, label=name),
            )
        )
    return tasks


def load_messages(path: Path) -> list[ChatMessage]:
    data = json.loads(path.read_text(encoding="utf-8"))
    messages: list[ChatMessage] = []
    for msg in data["messages"]:
        msg_id = str(msg["id"])
        text = str(msg["text"])
        messages.append(
            ChatMessage(
                msg_id=msg_id,
                channel=str(msg.get("channel") or data.get("channel") or ""),
                user=str(msg["user"]),
                ts=as_datetime(msg["ts"]),
                text=text,
                source_ref=SourceRef(system="slack", ref_id=msg_id, label=text[:80]),
            )
        )
    return messages


def load_github(path: Path) -> tuple[list[Commit], list[PullRequest]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    commits = [
        Commit(
            sha=str(row["sha"]),
            message=str(row["message"]),
            author=str(row["author"]),
            date=as_datetime(row["date"]),
            jira_keys=jira_keys(str(row["message"])),
            source_ref=SourceRef(
                system="github",
                ref_id=str(row["sha"])[:7],
                label=str(row["message"]),
            ),
        )
        for row in data["commits"]
    ]
    prs = [
        PullRequest(
            number=int(row["number"]),
            title=str(row["title"]),
            author=str(row["author"]),
            reviewers=list(row.get("reviewers") or []),
            state=str(row["state"]),
            created=as_datetime(row["created_at"]),
            merged_at=as_datetime(row["merged_at"]) if row.get("merged_at") else None,
            ci_status=str(row.get("ci_status") or ""),
            jira_keys=jira_keys(str(row["title"])),
            source_ref=SourceRef(
                system="github",
                ref_id=f"PR-{row['number']}",
                label=str(row["title"]),
            ),
        )
        for row in data["pull_requests"]
    ]
    return commits, prs


def minute_date_from_file(path: Path) -> date:
    match = re.search(r"minutes_(\d{4}-\d{2}-\d{2})_", path.name)
    if not match:
        raise ValueError(f"Cannot parse meeting date from {path.name}")
    return date.fromisoformat(match.group(1))


def load_minutes(minutes_dir: Path) -> list[MinuteItem]:
    kind_by_heading = {
        "Decisions": "decision",
        "Action Items": "action_item",
        "Questions": "question",
    }
    items: list[MinuteItem] = []
    for path in sorted(minutes_dir.glob("*.docx")):
        meeting_date = minute_date_from_file(path)
        current_kind: str | None = None
        counters = {"decision": 0, "action_item": 0, "question": 0}
        for paragraph in Document(path).paragraphs:
            text = paragraph.text.strip()
            if not text:
                continue
            if paragraph.style.name == "Heading 1" and text in kind_by_heading:
                current_kind = kind_by_heading[text]
                continue
            if paragraph.style.name == "List Bullet" and current_kind:
                counters[current_kind] += 1
                ref_id = f"{path.name}#{current_kind}-{counters[current_kind]}"
                items.append(
                    MinuteItem(
                        meeting_date=meeting_date,
                        kind=current_kind,
                        text=text,
                        source_ref=SourceRef(
                            system="minutes",
                            ref_id=ref_id,
                            label=text[:80],
                        ),
                    )
                )
    return items


def build_state() -> ProjectState:
    config = load_config()
    sample_dir = REPO_ROOT / "data" / "sample"
    github_commits, github_prs = load_github(sample_dir / "Sample_GitHub_Activity.json")
    return ProjectState(
        project_name=str(config["project"]),
        period_start=as_date(config["period"]["start"]),
        period_end=as_date(config["period"]["end"]),
        issues=load_issues(sample_dir / "Sample_Jira_Export.xlsx"),
        wbs_tasks=load_wbs(sample_dir / "Sample_WBS.xlsx"),
        messages=load_messages(sample_dir / "Sample_Slack_Messages.json"),
        commits=github_commits,
        prs=github_prs,
        minute_items=load_minutes(sample_dir / "minutes"),
        sprints=load_sprints(sample_dir / "Sample_Jira_Export.xlsx"),
    )


def main() -> None:
    state = build_state()
    out_path = REPO_ROOT / "data" / "sample" / "_projectstate.json"
    out_path.write_text(state.model_dump_json(indent=2), encoding="utf-8")

    counts = {
        "issues": len(state.issues),
        "wbs_tasks": len(state.wbs_tasks),
        "messages": len(state.messages),
        "commits": len(state.commits),
        "prs": len(state.prs),
        "minute_items": len(state.minute_items),
        "sprints": len(state.sprints),
    }
    print("ProjectState dumped to", out_path)
    print(" ".join(f"{key}={value}" for key, value in counts.items()))


if __name__ == "__main__":
    main()
